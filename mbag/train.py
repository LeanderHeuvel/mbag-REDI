# -*- coding: utf-8 -*-
"""
Created on Wed Sep 22 16:43:24 2021

@author: PathakS
"""

import os
import math
import torch
import datetime
import sys
import argparse

import numpy as np
import pandas as pd
import openpyxl as op
import torch.nn as nn
import torch.optim as optim

import matplotlib.pyplot as plt
from torchvision import transforms

import torch.nn.functional as F
from sklearn.metrics import confusion_matrix
from torch.utils.data import DataLoader
from openpyxl import Workbook

torch.autograd.set_detect_anomaly(True)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.dirname(SCRIPT_DIR))

from utils.pytorchtools import EarlyStopping
import utils.utils as utils
import mbag

def results_store_excel(correct_train,total_images_train,train_loss,correct_test,total_images_test,test_loss,epoch,conf_mat_train,conf_mat_test):
    avg_train_loss=train_loss/total_images_train
    avg_test_loss=test_loss/total_images_test
    accuracy_train=correct_train / total_images_train
    accuracy_test=correct_test / total_images_test
    recall_train=conf_mat_train[1,1]/sum(conf_mat_train[1,:])
    recall_test=conf_mat_test[1,1]/sum(conf_mat_test[1,:])
    speci_train=conf_mat_train[0,0]/sum(conf_mat_train[0,:])
    speci_test=conf_mat_test[0,0]/sum(conf_mat_test[0,:])
    lines=[epoch+1, avg_train_loss, accuracy_train, recall_train, speci_train, avg_test_loss, accuracy_test, recall_test, speci_test]
    out=open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    sheet1.append(lines)
    out.close()

def results_plot(df, file_name, base_path):
    plt.plot(pd.to_numeric(df['Epoch']),pd.to_numeric(df['Accuracy Val']),'-r',label='Accuracy Val')
    plt.plot(pd.to_numeric(df['Epoch']),pd.to_numeric(df['Accuracy Train']),'-b',label='Accuracy Train')
    plt.plot(pd.to_numeric(df['Epoch']),pd.to_numeric(df['Avg Loss Train']),'-g',label='Train Loss')
    plt.plot(pd.to_numeric(df['Epoch']),pd.to_numeric(df['Avg Loss Val']),'-y',label='Val Loss')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,pd.to_numeric(df.iloc[-1]['Epoch'])))
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy/Loss')
    plt.title(file_name)
    plt.savefig(base_path+'/multiview_mammogram/results/'+file_name+'.png')
    plt.show()

def conf_mat_create(predicted,true,correct,total_images,conf_mat):
    total_images+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat_batch=confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    conf_mat=conf_mat+conf_mat_batch
    return correct, total_images,conf_mat,conf_mat_batch

def save_model(model,optimizer,epoch):
    state = {'epoch': epoch+1,
             'state_dict': model.state_dict(),
             'optim_dict' : optimizer.state_dict()
            }
    torch.save(state,path_to_model)

def load_model(model,optimizer,path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    optimizer.load_state_dict(checkpoint['optim_dict'])
    epoch = checkpoint['epoch']
    return model,optimizer,epoch

def adaptive_learning_rate(optimizer, epoch, init_lr=0.001):
    """Sets the learning rate to the initial LR decayed by 0.2 every 10 epochs"""
    lr = init_lr * (0.2 ** (epoch // 10))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr
    
    return optimizer

def optimizer_fn():
    optimizer = optim.SGD(filter(lambda p: p.requires_grad, model.parameters()), lr=0.0001, momentum=0.9, weight_decay=0.0005)
    return optimizer

def loss_fn_softmax(weighted_cost_func, class_weights=None):
    if weighted_cost_func:
        criterion=nn.CrossEntropyLoss(class_weights)
    else:
        criterion=nn.CrossEntropyLoss()
    return criterion

def train(model,data_iterator_train,data_iterator_test,batches_train,batches_val,epochs):
    '''Training'''
    optimizer=optimizer_fn()
    early_stopping = EarlyStopping(path_to_model=path_to_model,patience=patience_epoch,verbose=True)
    if os.path.isfile(path_to_model):
        model,optimizer,start_epoch=load_model(model,optimizer,path_to_model)
        optimizer = adaptive_learning_rate(optimizer,start_epoch)
        print("start epoch:",start_epoch)
    else:
        start_epoch=0
    
    if weighted_cost_func:
        class_weights = utils.class_distribution_weightedloss(df_train)
    else:
        class_weights = None
    
    if activation=='softmax':
        lossfn = loss_fn_softmax(weighted_cost_func, class_weights)
    
    for epoch in range(start_epoch,epochs):
        optimizer=adaptive_learning_rate(optimizer,epoch)
        model.train()
        loss_train=0.0
        correct_train=0
        conf_mat_train=np.zeros((2,2))
        total_images_train=0
        batch_no=0
        loss_ar_train=[]
        for train_idx, train_batch, train_labels in data_iterator_train:
            train_batch = train_batch.to(device)
            train_labels = train_labels.to(device)
            train_labels=train_labels.view(-1)
            print("train batch:",train_batch.shape)
            output_batch = model(train_batch) # compute model output, loss and total train loss over one epoch
            
            if activation=='softmax':
                pred = output_batch.argmax(dim=1, keepdim=True)
                loss = lossfn(output_batch, train_labels)
            
            print("output batch shape:",output_batch.shape)
            loss_train+=(train_labels.size()[0]*loss.item())
            
            optimizer.zero_grad()  # clear previous gradients, compute gradients of all variables wrt loss
            loss.backward()
            optimizer.step() # performs updates using calculated gradients
            
            #performance metrics of training dataset
            correct_train,total_images_train,conf_mat_train,_=conf_mat_create(pred, train_labels, correct_train, total_images_train, conf_mat_train)
            batch_no=batch_no+1
            if batch_no:
                loss_ar_train.append(float(loss_train)/total_images_train)
            print('Train: Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}'.format(epoch+1, epochs, batch_no, batches_train, loss.item()))
        
        correct_test,total_images_test,loss_test,conf_mat_test=validation(model, data_iterator_test, epoch, batches_val)
        #print("total images in the whole training data for one epoch of training and test:",total_images_train,total_images_test)
        results_store_excel(correct_train,total_images_train,loss_train,correct_test,total_images_test,loss_test,epoch, conf_mat_train, conf_mat_test)
        valid_loss=loss_test/total_images_test
        
        # if epoch==start_epoch:
        #     loss_np_train=np.array([loss_ar_train])
        #     # loss_np_val=np.array([loss_ar_val])
        # else:
        #     loss_np_train=np.append(loss_np_train,[np.array(loss_ar_train)],axis=0)
            # loss_np_val=np.append(loss_np_val,[np.array(loss_ar_val)],axis=0)
        # early_stopping needs the validation loss to check if it has decresed, 
        # and if it has, it will make a checkpoint of the current model
        early_stopping(valid_loss,model,optimizer,epoch,conf_mat_train,conf_mat_test)
        if early_stopping.early_stop:
            print("Early stopping",epoch+1)
            break
        
    
    sheet2.append([0,1])
    for row in early_stopping.conf_mat_train_best.tolist():
        sheet2.append(row)
    sheet2.append([0,1])
    for row in early_stopping.conf_mat_test_best.tolist():
        sheet2.append(row)
    np_out=open('train_loss_batches.npy','wb')
    # np.save(np_out,loss_np_train)
    np_out.close()
    np_out2=open('val_loss_batches.npy','wb')
    # np.save(np_out2,loss_np_val)
    np_out2.close()
    print('Finished Training')
    
def validation(model, data_iterator_val, epoch, batches_val):
    """Validation"""
    model.eval()
    total_images=0
    val_loss = 0
    correct = 0
    s=0
    batch_val_no=0
    conf_mat_test=np.zeros((2,2))
    if weighted_cost_func:
        class_weights_val = utils.class_distribution_poswt(df_val)
    else:
        class_weights_val = None
    
    if activation=='softmax':
        lossfn1 = loss_fn_softmax(weighted_cost_func, class_weights_val)
    #loss_ar_val=[]
    with torch.no_grad():   
        for val_idx, val_batch, val_labels in data_iterator_val:
            val_batch, val_labels=val_batch.to(device), val_labels.to(device)
            val_labels=val_labels.view(-1)#.float()
            output_val = model(val_batch)
            if activation=='softmax':
                val_pred = output_val.argmax(dim=1, keepdim=True)
                loss1 = lossfn1(output_val, val_labels).item()
            
            s=s+val_labels.shape[0]    
            val_loss += val_labels.size()[0]*loss1 # sum up batch loss
            correct,total_images,conf_mat_test,_=conf_mat_create(val_pred,val_labels,correct,total_images,conf_mat_test)
            batch_val_no+=1
            #if batch_val_no:
            #    loss_ar_val.append(float(val_loss)/s) 
            print('Val: Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}'.format(epoch+1, max_epochs, batch_val_no, batches_val, loss1))
    
    print("conf_mat_test:",conf_mat_test)
    print("total_images:",total_images)
    print("s:",s)
    print('\nTest set: total val loss: {:.4f}, Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%), Epoch:{}\n'.format(
        val_loss, val_loss/total_images, correct, total_images,
        100. * correct / total_images,epoch+1))
    return correct,total_images,val_loss,conf_mat_test#,n_p_ratio_val

def test(model, data_iterator_test, batches_test):
    """Testing"""
    model.eval()
    total_images=0
    test_loss = 0
    correct = 0
    s=0
    batch_test_no=0
    conf_mat_test=np.zeros((2,2))
    lossfn1 = loss_fn_softmax(False)
    for test_idx, test_batch, test_labels in data_iterator_test:
        test_batch, test_labels=test_batch.to(device), test_labels.to(device)
        test_labels=test_labels.view(-1)
        output_test = model(test_batch)
        if activation=='softmax':
            test_pred = output_test.argmax(dim=1, keepdim=True)
        
        if batch_test_no==0:
            test_pred_all=test_pred
            test_labels_all=test_labels
            print(output_test.data.shape)
            if activation=='softmax':
                output_all_ten=F.softmax(output_test.data,dim=1)
                output_all_ten=output_all_ten[:,1]
        else:
            test_pred_all=torch.cat((test_pred_all,test_pred),dim=0)
            test_labels_all=torch.cat((test_labels_all,test_labels),dim=0)
            if activation=='softmax':
                output_all_ten=torch.cat((output_all_ten,F.softmax(output_test.data,dim=1)[:,1]),dim=0)
        
        loss1=lossfn1(output_test, test_labels).item()
        test_loss += test_labels.size()[0]*loss1 # sum up batch loss
        correct, total_images, conf_mat_test, _=conf_mat_create(test_pred,test_labels,correct,total_images,conf_mat_test)
        batch_test_no+=1
        s=s+test_batch.shape[0]
        print ('Test: Step [{}/{}], Loss: {:.4f}'.format(batch_test_no, batches_test, loss1))
    
    print("conf_mat_test:",conf_mat_test)
    print("total_images:",total_images)
    print("s:",s)
    print('\nTest set: total val loss: {:.4f}, Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%) \n'.format(
        test_loss, test_loss/total_images, correct, total_images, 100. * correct / total_images))
    
    sheet3.append([0,1])
    for row in conf_mat_test.tolist():
        sheet3.append(row)
    
    per_model_metrics = utils.performance_metrics(conf_mat_test,test_labels_all.cpu().numpy(),test_pred_all.cpu().numpy(), output_all_ten.cpu().numpy())
    print(per_model_metrics)
    sheet3.append(['Precision','Recall','Specificity','F1','Acc','Bal_Acc','Cohens Kappa','AUC'])
    sheet3.append(per_model_metrics)

def analyze_test_samples(model, data_iterator_test, batches_test):
    """Evaluating test samples"""
    model.eval()
    total_images=0
    test_loss = 0
    correct = 0
    s=0
    batch_test_no=0
    lossfn1 = loss_fn_softmax(False)
    sheet5.append(['ImgName','AbnormalityType','Groundtruth','Prediction'])
    for test_idx, test_batch, test_labels,img_name,abnormality_type  in data_iterator_test:
        # test_labels, img_name, abnormality_type = test_labels
        test_batch, test_labels=test_batch.to(device), test_labels.to(device)
        test_labels=test_labels.view(-1)
        output_test = model(test_batch)
        if activation=='softmax':
            test_pred = output_test.argmax(dim=1, keepdim=True)
        
        if batch_test_no==0:
            test_pred_all=test_pred
            test_labels_all=test_labels
            print(output_test.data.shape)
            if activation=='softmax':
                output_all_ten=F.softmax(output_test.data,dim=1)
                output_all_ten=output_all_ten[:,1]
        else:
            test_pred_all=torch.cat((test_pred_all,test_pred),dim=0)
            test_labels_all=torch.cat((test_labels_all,test_labels),dim=0)
            if activation=='softmax':
                output_all_ten=torch.cat((output_all_ten,F.softmax(output_test.data,dim=1)[:,1]),dim=0)
        
        loss1=lossfn1(output_test, test_labels).item()
        test_loss += test_labels.size()[0]*loss1 # sum up batch loss
        
        # correct, total_images, conf_mat_test, _=conf_mat_create(test_pred,test_labels,correct,total_images,conf_mat_test)
        batch_test_no+=1
        s=s+test_batch.shape[0]
        results = [str(img_name[0]), str(abnormality_type[0]), str(test_labels.cpu().numpy()[0]), str(test_pred.cpu().numpy()[0])]
        sheet5.append(results)

        # print ('Test: Step [{}/{}], Loss: {:.4f}'.format(batch_test_no, batches_test, loss1))

    # sheet3.append([0,1])
    # for row in conf_mat_test.tolist():
    #     sheet3.append(row)
    
    # per_model_metrics = utils.performance_metrics(conf_mat_test,test_labels_all.cpu().numpy(),test_pred_all.cpu().numpy(), output_all_ten.cpu().numpy())
    
    # print(per_model_metrics)
    # sheet5.append(['Precision','Recall','Specificity','F1','Acc','Bal_Acc','Cohens Kappa','AUC'])
    # sheet5.append(per_model_metrics)

    pass
    
if __name__=='__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--image_size", help="size of square image with height = width, ", type=int, default=1300)
    parser.add_argument("--patch_size", help="image patch size of the model, ", type=int, default=9)
    parser.add_argument("--train", help="image patch size of the model, ", type=int, default=1)
    args = parser.parse_args()
    training_required = args.train
    image_size = args.image_size
    patch_size = args.patch_size
    begin_time = datetime.datetime.now()
    #Initialization    
    modality='MG'
    num_classes=2
    patience_epoch=20
    groundtruth_dic={'benign':0,'malignant':1}
    classes=[0,1]
    max_epochs=100
    count=0
    count1=0
    acc_num_list_final=[]
    groundtruth_list=[]
    acc_num_firstsubset=[]
    use_pretrained=True
    batch_size=1
    num_workers=4
    mean=[0.5,0.5,0.5]
    std_dev=[0.5,0.5,0.5]
    rand_seed=8
    
    #some settings for the type of model I am training
    activation='softmax' #Kim et al:softmax; 
    attention='none'
    milpooling='average' #Kim et al:average; 
    weighted_cost_func=False #for class imbalance handling, set as True otherwise False
    model_type='single_pipeline' 
    flipimage = True
    
    torch.manual_seed(8)
    torch.cuda.manual_seed(8)
    np.random.seed(8)
    
    # CUDA for PyTorch
    use_cuda = torch.cuda.is_available()
    device = torch.device("cuda:0" if use_cuda else "cpu")
    print(device)
    
    #Output file names
    base_path = os.path.abspath(os.getcwd())+"/mbag-REDI"
    file_name="sota_singlepipeline_MILpoolingavg_softmax_variableview_batch15_pe10_oldresult_run2_bagnet_"+str(patch_size)+'_'+modality #name of the output files that will be created
    path_to_model = base_path+"/multiview_mammogram/models/"+file_name+".tar" #name of the folder where to save the models
    path_to_results = base_path+"/multiview_mammogram/results/"+file_name+".xlsx"
    path_to_results_text = base_path+"/multiview_mammogram/results/"+file_name+".txt"
    path_to_log_file = base_path+"/multiview_mammogram/results/"+file_name+"_log"+".txt"
    if not os.path.exists(base_path+'/multiview_mammogram/models'):
        os.mkdir(base_path+'/multiview_mammogram/models')
    if not os.path.exists(base_path+'/multiview_mammogram/results'):
        os.mkdir(base_path+'/multiview_mammogram/results')
    
    #input file names
    csv_file_modality=base_path+'/MG_training_files_cbis-ddsm_singleinstance_groundtruth_adapted.csv' #name of the file which contains path to the images and other information of the images. 
    df_modality=pd.read_csv(csv_file_modality, sep=';')
    #df_modality=df_modality[~df_modality['StudyInstanceUID'].isnull()]
    #print("the original df modality shape:",df_modality.shape)
    #df_modality=df_modality[~df_modality['Views'].isnull()]
    #print("df modality no null:",df_modality.shape)
    
    #View distribution and creates a new file with the view names mentioned and saves in input_data
    #utils.views_distribution(df_modality)
    #input('wait')
    
    #new data file after adding view information to the file and taking instances with exactly 4 views
    #df_modality=df_modality[:200]
    #df_modality1=df_modality[df_modality['Views'].str.split('+').str.len()==4.0]
    #print("df_modality 4 views:", df_modality1.shape)
    
    #Train-val-test split
    df_train, df_val, df_test = utils.stratifiedgroupsplit(df_modality, rand_seed)
    total_instances=df_modality.shape[0]
    
    df_train = df_train.reset_index()
    df_val = df_val.reset_index()
    df_test = df_test.reset_index()
    
    total_instances=df_modality.shape[0]
    train_instances=df_train.shape[0]
    val_instances=df_val.shape[0]
    test_instances=df_test.shape[0]
    print(utils.stratified_class_count(df_train))
    print(utils.stratified_class_count(df_val))
    print(utils.stratified_class_count(df_test))
    print("training instances:", train_instances)
    print("Validation instances:", val_instances)
    print("Test instances:",test_instances)
    
    # set file path
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        sheet1 = wb['epoch training']
        sheet2 = wb['confusion matrix train_val']
        sheet3 = wb['confusion matrix test']
        sheet4 = wb['metrics view wise']
        sheet5 = wb['eval per sample']
    else:
        wb=Workbook()
        sheet1=wb.active
        header=['Epoch','Avg Loss Train','Accuracy Train','Recall Train','Specificity Train','Avg Loss Val','Accuracy Val','Recall Val','Specificity Val']
        sheet1 = wb.create_sheet('epoch training')
        sheet1.append(header)
        sheet2 = wb.create_sheet('confusion matrix train_val')
        sheet3 = wb.create_sheet('confusion matrix test') 
        sheet4 = wb.create_sheet('metrics view wise')
        sheet5 = wb.create_sheet('eval per sample')
    if patch_size==9:
        model = mbag.bagnet9_18(num_classes=2) #change this line to resnet18,resnet50 or bagnet() whatever you want to use
    if patch_size==17:
        model = mbag.bagnet17_18(num_classes=2)
    if patch_size==33:
        model = mbag.bagnet33_18(num_classes=2)
    model= nn.DataParallel(model)
    model.to(device)
    pytorch_total_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    
    preprocess_train=utils.data_augmentation_train(mean,std_dev, image_size)
    
    preprocess_val = transforms.Compose([
        transforms.Resize((image_size,image_size)), #original was 1600 by 1600
        transforms.ToTensor(),
        transforms.Normalize(mean=mean, std=std_dev)
    ])
    
    dataset_gen_train = utils.BreastCancerDataset_generator(df_train,modality,flipimage, preprocess_train)
    dataloader_train = DataLoader(dataset_gen_train, batch_size=batch_size, shuffle=True, num_workers=num_workers , collate_fn=utils.MyCollate)   
    
    dataset_gen_val = utils.BreastCancerDataset_generator(df_val,modality,flipimage, preprocess_val)
    dataloader_val = DataLoader(dataset_gen_val, batch_size=batch_size, shuffle=False, num_workers=num_workers, collate_fn=utils.MyCollate)
    
    dataset_gen_test = utils.BreastCancerDataset_generator(df_test,modality,flipimage,preprocess_val)
    dataloader_test = DataLoader(dataset_gen_test, batch_size=batch_size, shuffle=False, num_workers=num_workers, collate_fn=utils.MyCollate)

    dataset_gen_test_evaluate = utils.BreastCancerDataset_generator(df_test,modality,flipimage,preprocess_val, get_image_name=True)
    dataloader_evaluate = DataLoader(dataset_gen_test_evaluate, batch_size=batch_size, shuffle=False, num_workers=num_workers, collate_fn=utils.MyCollate)

    batches_train=int(math.ceil(train_instances/batch_size))
    batches_val=int(math.ceil(val_instances/batch_size))
    batches_test=int(math.ceil(test_instances/batch_size))
    
    #training and validation
    if training_required == 1:
        train(model, dataloader_train, dataloader_val, batches_train, batches_val, max_epochs)
    
    optimizer = optimizer_fn()
    path_to_trained_model=path_to_model
    model,  optimizer, epoch_idx = load_model(model, optimizer, path_to_trained_model)
    if training_required == 2:
        ''' evaluation per sample '''
        analyze_test_samples(model,dataloader_evaluate, batches_test)
    else:
        test(model, dataloader_test, batches_test)
    wb.save(path_to_results)
            
    #plot the training and validation loss and accuracy
    df=pd.read_excel(path_to_results, sheet_name='epoch training')
    df_text = pd.read_csv(path_to_results_text, sep=" ", header=None)
    df_text.columns = ['Epoch','Avg Loss Train','Accuracy Train','Recall Train','Specificity Train','Avg Loss Val','Accuracy Val','Recall Val','Specificity Val']
    for column in df_text:
        df_text[column] = df_text[column].str.extract(r'(\d+[.\d]*)')
    results_plot(df_text,file_name, base_path)
    print("End time:",datetime.datetime.now())
    print("Execution time:",datetime.datetime.now() - begin_time)

    f = open(path_to_log_file,'w')
    f.write("Model parameters:"+str(pytorch_total_params/math.pow(10,6))+'\n')
    f.write("Start time:"+str(begin_time)+'\n')
    f.write("End time:"+str(datetime.datetime.now())+'\n')
    f.write("Execution time:"+str(datetime.datetime.now() - begin_time)+'\n')
    f.close()


    
